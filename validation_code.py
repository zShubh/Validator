import streamlit as st
import pandas as pd
import io
import numpy as np
from datetime import datetime
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO
import tempfile
from copy import copy
from datetime import datetime
import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font
import re
import traceback
from copy import copy



# Validation Report Analyzer functions
def analyze_validation_report(file_content, file_type):
    """
    Analyzes a validation report from a CSV or Excel file, handling multiple sheets in Excel.

    Args:
        file_content (bytes): The content of the uploaded file.
        file_type (str): The type of the uploaded file ('csv' or 'xlsx').

    Returns:
        dict: A dictionary containing the analysis results with these keys:
            - 'comparison_df': DataFrame containing PBI and Cognos column names, diffs, Unique ID and Presence.
            - 'error': A string describing any error that occurred, or None if no error.
            - 'presence_analysis': dict containing counts of 'PBI' and 'Cognos' in 'Presence' column, and differing IDs
    """
    
    try:
        if file_type == 'csv':
            # Read the CSV data
            df = pd.read_csv(io.BytesIO(file_content))
        elif file_type == 'xlsx':
            # Read the Excel file, targeting the "Validation_Report" sheet.  Crucially,
            # we do NOT return here if the sheet is not found.  Instead, we let the
            # rest of the analysis proceed, which is what the user wants.  We DO
            # still need to catch the error, so the 'try' block remains.
            try:
                df = pd.read_excel(io.BytesIO(file_content), sheet_name="Validation_Report")
            except KeyError:
                # Instead of returning, set df to an empty DataFrame and continue.
                # This ensures the rest of the code runs. We also set a custom attribute
                # on the DataFrame to indicate that the sheet was not found.
                df = pd.DataFrame()
                df.sheet_not_found= True
        else:
            return {'error': f"Error: Unsupported file type: {file_type}"}

    except Exception as e:
        return {'error': f'Error reading file: {e}'}

    # 1. Identify Matching Prefixes
    pbi_cols = [col for col in df.columns if col.endswith('_PBI')]
    cognos_cols = [col for col in df.columns if col.endswith('_Cognos')]
    diff_cols = [col for col in df.columns if col.endswith('_Diff')]
    id_col = 'unique_key'
    presence_col = 'presence'

    # 2. Extract Common Prefixes
    pbi_prefixes = [col.replace('_PBI', '') for col in pbi_cols]
    cognos_prefixes = [col.replace('_Cognos', '') for col in cognos_cols]
    diff_prefixes = [col.replace('_Diff', '') for col in diff_cols]
    common_prefixes = list(set(pbi_prefixes) & set(cognos_prefixes) & set(diff_prefixes))

    if not common_prefixes:
        return {'error': "Error: No common column prefixes found in the data."}

    # 3. Create Comparison Dataframe
    comparison_df = pd.DataFrame()
    comparison_df['Unique ID'] = df[id_col]
    comparison_df['Presence'] = df[presence_col] 

    for prefix in common_prefixes:
        comparison_df[f'{prefix}_PBI'] = df[f'{prefix}_PBI']
        comparison_df[f'{prefix}_Cognos'] = df[f'{prefix}_Cognos']
        comparison_df[f'{prefix}_Diff'] = df[f'{prefix}_Diff']

    # 4. Analyze 'Presence' Column
    presence_analysis = {}
    if presence_col in df:
        presence_counts = df[presence_col].value_counts()
        presence_analysis['pbi_present_count'] = presence_counts.get('Present in PBI', 0)
        presence_analysis['cognos_present_count'] = presence_counts.get('Present in Cognos', 0)

        # Find Unique IDs where Presence is 'Present in PBI' or 'Present in Cognos'
        pbi_present_ids = df[df[presence_col] == 'Present in PBI'][id_col].tolist()
        cognos_present_ids = df[df[presence_col] == 'Present in Cognos'][id_col].tolist()
        presence_analysis['pbi_present_ids'] = sorted(pbi_present_ids)
        presence_analysis['cognos_present_ids'] = sorted(cognos_present_ids)
    else:
        presence_analysis['pbi_present_count'] = 0
        presence_analysis['cognos_present_count'] = 0
        presence_analysis['pbi_present_ids'] = []
        presence_analysis['cognos_present_ids'] = []

    return {
        'comparison_df': comparison_df,
        'error': None,
        'presence_analysis': presence_analysis,
        'sheet_not_found': getattr(df, 'sheet_not_found', False) # Get custom attribute, default to False
    }


# Main functions for Validation Report Generator
# Function to strip leading zeros and convert to numeric if applicable
def strip_leading_zeros(val):
    try:
        if isinstance(val, str):
            val = val.strip().replace(',', '')
            if val.replace('.', '', 1).isdigit() or \
               (val.startswith('-') and val[1:].replace('.', '', 1).isdigit()):
                return float(val)
        return val
    except:
        return val

# Apply numeric cleaning only on likely numeric columns
def convert_possible_numeric(df):
    for col in df.columns:
        df[col] = df[col].apply(strip_leading_zeros)
        # Try numeric conversion first
        try:
            df[col] = pd.to_numeric(df[col], errors='ignore')
        except:
            pass

        # Then try datetime conversion if it's still an object type
        if df[col].dtype == 'object':
            try:
                # Attempt to parse dates, inferring format
                converted_dates = pd.to_datetime(df[col], errors='coerce')
                # Check if a significant portion of the column was converted to datetime
                non_na_count = df[col].notna().sum()
                successful_conversion_count = converted_dates.notna().sum()
                if successful_conversion_count > 0.8 * non_na_count:  # Convert if at least 80% are likely dates
                    df[col] = converted_dates
            except:
                pass
    return df

# Define the checklist data as a DataFrame
checklist_data = {
    "S.No": range(1, 18),
    "Checklist": [
        "Database & Warehouse is parameterized (In case of DESQL Reports)",
        "All the columns of Cognos replicated in PBI (No extra columns)",
        "All the filters of Cognos replicated in PBI",
        "Filters working as expected (single/multi select as usual)",
        "Column names matching with Cognos",
        "Currency symbols to be replicated",
        "Filters need to be aligned vertically/horizontally",
        "Report Name & Package name to be written",
        "Entire model to be refreshed before publishing to PBI service",
        "Date Last refreshed to be removed from filter/table",
        "Table's column header to be bold",
        "Table style to not have grey bars",
        "Pre-applied filters while generating validation report?",
        "Dateformat to beYYYY-MM-DD [hh:mm:ss] in refresh date as well",
        "Sorting is replicated",
        "Filter pane to be hidden before publishing to PBI service",
        "Mentioned the exception in our validation document like numbers/columns/values mismatch (if any)"
    ],
    "Status - Level1": ["" for _ in range(17)],
    "Status - Level2": ["" for _ in range(17)]
}
checklist_df = pd.DataFrame(checklist_data)

def generate_validation_report(cognos_df, pbi_df, dimension_columns):
    dims = dimension_columns

    cognos_df[dims] = cognos_df[dims].fillna('NAN')
    pbi_df[dims] = pbi_df[dims].fillna('NAN')

    cognos_measures = [col for col in cognos_df.columns if col not in dims and np.issubdtype(cognos_df[col].dtype, np.number)]
    pbi_measures = [col for col in pbi_df.columns if col not in dims and np.issubdtype(pbi_df[col].dtype, np.number)]
    all_measures = list(set(cognos_measures) & set(pbi_measures))

    cognos_agg = cognos_df.groupby(dims)[all_measures].sum().reset_index()
    pbi_agg = pbi_df.groupby(dims)[all_measures].sum().reset_index()

    cognos_agg['unique_key'] = cognos_agg[dims].astype(str).agg('-'.join, axis=1).str.upper()
    pbi_agg['unique_key'] = pbi_agg[dims].astype(str).agg('-'.join, axis=1).str.upper()

    cognos_agg = cognos_agg[['unique_key'] + [col for col in cognos_agg.columns if col != 'unique_key']]
    pbi_agg = pbi_agg[['unique_key'] + [col for col in pbi_agg.columns if col != 'unique_key']]

    validation_report = pd.DataFrame({'unique_key': list(set(cognos_agg['unique_key']) | set(pbi_agg['unique_key']))})

    for dim in dims:
        validation_report[dim] = validation_report['unique_key'].map(dict(zip(cognos_agg['unique_key'], cognos_agg[dim])))
        validation_report[dim].fillna(validation_report['unique_key'].map(dict(zip(pbi_agg['unique_key'], pbi_agg[dim]))), inplace=True)

    validation_report['presence'] = validation_report['unique_key'].apply(
        lambda key: 'Present in Both' if key in cognos_agg['unique_key'].values and key in pbi_agg['unique_key'].values
        else ('Present in Cognos' if key in cognos_agg['unique_key'].values
              else 'Present in PBI')
    )

    for measure in all_measures:
        validation_report[f'{measure}_Cognos'] = validation_report['unique_key'].map(dict(zip(cognos_agg['unique_key'], cognos_agg[measure])))
        validation_report[f'{measure}_PBI'] = validation_report['unique_key'].map(dict(zip(pbi_agg['unique_key'], pbi_agg[measure])))

        validation_report[f'{measure}_Diff'] = validation_report[f'{measure}_PBI'].fillna(0) - validation_report[f'{measure}_Cognos'].fillna(0)

    column_order = ['unique_key'] + dims + ['presence'] + [col for measure in all_measures for col in
                                                            [f'{measure}_Cognos', f'{measure}_PBI', f'{measure}_Diff']]
    validation_report = validation_report[column_order]

    return validation_report, cognos_agg, pbi_agg

def column_checklist(cognos_df, pbi_df):
    cognos_columns = cognos_df.columns.tolist()
    pbi_columns = pbi_df.columns.tolist()

    checklist_df = pd.DataFrame({
        'Cognos Columns': cognos_columns + [''] * (max(len(pbi_columns), len(cognos_columns)) - len(cognos_columns)),
        'PowerBI Columns': pbi_columns + [''] * (max(len(pbi_columns), len(cognos_columns)) - len(pbi_columns))
    })

    checklist_df['Match'] = checklist_df.apply(lambda row: row['Cognos Columns'] == row['PowerBI Columns'], axis=1)

    return checklist_df

def generate_diff_checker(validation_report):
    diff_columns = [col for col in validation_report.columns if col.endswith('_Diff')]

    diff_checker = pd.DataFrame({
        'Diff Column Name': diff_columns,
        'Sum of Difference': [validation_report[col].sum() for col in diff_columns]
    })

    presence_summary = {
        'Diff Column Name': 'All rows present in both',
        'Sum of Difference': 'Yes' if all(validation_report['presence'] == 'Present in Both') else 'No'
    }
    diff_checker = pd.concat([diff_checker, pd.DataFrame([presence_summary])], ignore_index=True)

    return diff_checker

# Dry Run functions
def load_file(uploaded_file):
    if uploaded_file.name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    else:
        return pd.read_excel(uploaded_file)

def dry_run(file1, file2):
    """
    Performs a dry run comparison between two uploaded files.  It automatically
    identifies and compares numeric columns.

    Args:
        file1 (UploadedFile): The first file to compare.
        file2 (UploadedFile): The second file to compare.
    """
    if file1 is not None and file2 is not None:
        try:
            df1 = load_file(file1)
            df2 = load_file(file2)

            st.subheader("📏 Row Counts")
            col1, col2 = st.columns(2)
            col1.metric("cognos_file Rows", len(df1))
            col2.metric("pbi_file Rows", len(df2))

            # Show common columns
            st.subheader("🧾 Common Columns for Comparison")
            common_cols = list(set(df1.columns) & set(df2.columns))
            st.write(common_cols)

            # **Automatic Column Selection:**
            # Identify common columns that are numeric in both DataFrames
            numeric_cols1 = df1.select_dtypes(include='number').columns
            numeric_cols2 = df2.select_dtypes(include='number').columns
            selected_columns = list(set(numeric_cols1) & set(numeric_cols2))
            st.write(f"Automatically selected numeric columns: {selected_columns}") #show selected columns

            if selected_columns:
                df1_numeric = df1[selected_columns].copy()  # Create a copy to avoid modifying the original DataFrame
                df2_numeric = df2[selected_columns].copy()  # Create a copy
                
                # Convert to numeric, errors='coerce' will turn non-numeric to NaN
                for col in selected_columns:
                    df1_numeric[col] = pd.to_numeric(df1_numeric[col], errors='coerce')
                    df2_numeric[col] = pd.to_numeric(df2_numeric[col], errors='coerce')

                sum1 = df1_numeric.sum(skipna=True)  # Use skipna=True to ignore NaN values
                sum2 = df2_numeric.sum(skipna=True)
                

                percentage_diff = (abs(sum2 - sum1) / sum1) * 100

                comparison = pd.DataFrame({
                    "File 1 Sum": sum1,
                    "File 2 Sum": sum2,
                    "Absolute Difference": (sum1 - sum2).abs(),
                    "% Difference (w.r.t File 1)": percentage_diff.round(2)
                })

                st.subheader("📊 Column-wise Sum Comparison")
                st.dataframe(comparison)

            else:
                st.info("No common numeric columns found to compare.")

        except Exception as e:
            st.error(f"Error: {e}")

    elif file1 is not None or file2 is not None:
        st.warning("Please upload both files to compare.")



def main():
    selected_column_name = []
    st.set_page_config(page_title="Validation Report Tool", layout="wide")
    st.markdown("""
    <style>
        .big-font {
            font-size:24px !important;
            font-weight: bold;
        }
    </style>
    """, unsafe_allow_html=True)

    st.title("📊 Validation Report Tool 🚀")

    option = st.sidebar.radio("Choose Operation", ["Generate Validation Report","Excel Combiner"])

    if option == "Generate Validation Report":
        header_text = """
    <p class="big-font">
        📝 Important Assumptions:
    </p>
    <ol>
        <li> 📂 Upload two independent Excel files: "Cognos Data" and "PBI Data".</li>
        <li> 🏷️ Make sure the column names are similar in both sheets.</li>
        <li> 📌 Select columns to build the ID part.</li>
        <li>If the Title and Sub-title in Cognos are different, replicate the Sub-title with font size 16 and bold and color - black.</li>
        <li>Replicate the "$" symbol if it is present in the Cognos report.</li>
        <li>If the "$" symbol is present and the value is negative (negative currency numbers), we are using parentheses by default. However, if negative non-currency numbers in Cognos are shown in parentheses, we simply display the default negative numbers without replicating the parentheses.</li>
        <li>Matrix validation should be done in Excel: Sheet 1: Checklist, Sheet 2: Screenshots of both Power BI and Cognos.</li>
        <li>Please create a sheet and add Screenshots of PBI and Cognos in every Validation report.</li>
    </ol>
    """
        header_text = header_text.replace("<li>If the Title and Sub-title in Cognos are different, replicate the Sub-title with font size 16 and bold and color - black.</li>", "<li>🎨 If the Title and Sub-title in Cognos are different, replicate the Sub-title with font size 16 and bold and color - black.</li>")
        header_text = header_text.replace("<li>Replicate the \"$\" symbol if it is present in the Cognos report.</li>", "<li>💲 Replicate the \"$\" symbol if it is present in the Cognos report.</li>")
        header_text = header_text.replace("<li>If the \"$\" symbol is present and the value is negative (negative currency numbers), we are using parentheses by default. However, if negative non-currency numbers in Cognos are shown in parentheses, we simply display the default negative numbers without replicating the parentheses.</li>", "<li>📉 If the \"$\" symbol is present and the value is negative (negative currency numbers), we are using parentheses by default. However, if negative non-currency numbers in Cognos are shown in parentheses, we simply display the default negative numbers without replicating the parentheses.</li>")
        header_text = header_text.replace("<li>Matrix validation should be done in Excel: Sheet 1: Checklist, Sheet 2: Screenshots of both Power BI and Cognos.</li>", "<li>📊 Matrix validation should be done in Excel: Sheet 1: Checklist, Sheet 2: Screenshots of both Power BI and Cognos.</li>")
        header_text = header_text.replace("<li>Please create a sheet and add Screenshots of PBI and Cognos in every Validation report.</li>", "<li> 📸 Please create a sheet and add Screenshots of PBI and Cognos in every Validation report.</li>")

        st.markdown(header_text, unsafe_allow_html=True)

        model_name = st.text_input("Enter the model name:")
        report_name = st.text_input("Enter the report name:")
        
        
        cognos_file = st.file_uploader("Upload Cognos Data Excel file 📈", type="xlsx", key="cognos_upload")
        pbi_file = st.file_uploader("Upload PBI Data Excel file 📉", type="xlsx", key="pbi_upload")
        dry_run(cognos_file, pbi_file)

        if cognos_file is not None and pbi_file is not None:
            try:
                cognos_df = pd.read_excel(cognos_file)
                pbi_df = pd.read_excel(pbi_file)

                # Convert numeric-like strings
                cognos_df = convert_possible_numeric(cognos_df)
                pbi_df = convert_possible_numeric(pbi_df)

                # Standardize text columns
                cognos_df = cognos_df.apply(lambda x: x.str.upper().str.strip() if x.dtype == "object" else x)
                pbi_df = pbi_df.apply(lambda x: x.str.upper().str.strip() if x.dtype == "object" else x)

                option_data = st.radio("Select Option 🛠️", ["Data Present 📊", "Only Column Names Present 🏷️"])

                if option_data == "Only Column Names Present 🏷️":
                    column_checklist_df = column_checklist(cognos_df, pbi_df)

                    st.subheader("Column Checklist Preview 📋")
                    st.dataframe(column_checklist_df)

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        checklist_df.to_excel(writer, sheet_name='Checklist', index=False)
                        cognos_df.to_excel(writer, sheet_name='Cognos', index=False)
                        pbi_df.to_excel(writer, sheet_name='PBI', index=False)
                        column_checklist_df.to_excel(writer, sheet_name='Column Checklist', index=False)
                        pd.DataFrame().to_excel(writer, sheet_name='Cognos SS', index=False)
                        pd.DataFrame().to_excel(writer, sheet_name='PBI SS', index=False)

                    output.seek(0)
                    today_date = datetime.today().strftime('%Y-%m-%d')
                    dynamic_filename = f"{model_name}_{report_name}_ColumnCheck_Report_{today_date}.xlsx" if model_name and report_name else f"ColumnCheck_Report_{today_date}.xlsx"

                    st.download_button(
                        label="Download Column Check Excel Report ⬇️",
                        data=output,
                        file_name=dynamic_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.info("Once downloaded, please browse to your downloads folder to access the report.")

                elif option_data == "Data Present 📊":
                    common_columns = list(set(cognos_df.columns) & set(pbi_df.columns))
                    selected_columns = st.multiselect("Select columns to build the ID part: 🔑", common_columns)
                    if selected_columns:
                        validation_report, cognos_agg, pbi_agg = generate_validation_report(cognos_df, pbi_df, selected_columns)
                        column_checklist_df = column_checklist(cognos_df, pbi_df)
                        diff_checker_df = generate_diff_checker(validation_report)

                        st.subheader("Validation Report Preview 📈📉")
                        st.dataframe(validation_report)

                        # Checklist input section
                        st.subheader("Checklist Status 📝")
                        for i, row in checklist_df.iterrows():
                            checklist_df.loc[i, 'Status - Level1'] = st.text_input(f"Status - Level1 for {row['Checklist']} ✅", key=f"level1_{i}")
                            checklist_df.loc[i, 'Status - Level2'] = "" #Only Level 1 input.

                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            checklist_df.to_excel(writer, sheet_name='Checklist', index=False)
                            cognos_agg.to_excel(writer, sheet_name='Cognos', index=False)
                            pbi_agg.to_excel(writer, sheet_name='PBI', index=False)
                            validation_report.to_excel(writer, sheet_name='Validation_Report', index=False)
                            column_checklist_df.to_excel(writer, sheet_name='Column Checklist', index=False)
                            diff_checker_df.to_excel(writer, sheet_name='Diff Checker', index=False)
                            pd.DataFrame().to_excel(writer, sheet_name='Cognos SS', index=False)
                            pd.DataFrame().to_excel(writer, sheet_name='PBI SS', index=False)

                        output.seek(0)
                        today_date = datetime.today().strftime('%Y-%m-%d')
                        dynamic_filename = f"{model_name}_{report_name}_ValidationReport_{today_date}.xlsx" if model_name and report_name else f"ValidationReport_{today_date}.xlsx"

                        st.download_button(
                            label="Download Excel Report ⬇️",
                            data=output,
                            file_name=dynamic_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.info("Once downloaded, please browse to your downloads folder to access the report.")

                    else:
                        st.warning("Please select at least one column to build the ID part. ⚠️")
                
                

            except Exception as e:
                st.error(f"An error occurred: {str(e)} ❌")
    
            st.header("Validation Report Analyzer")
            uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=["csv", "xlsx"])

            if uploaded_file is not None:
                file_content = uploaded_file.getvalue()
                file_type = uploaded_file.name.split('.')[-1].lower()

                # Analyze the data
                results = analyze_validation_report(file_content, file_type)

                # Print the results
                if results['error']:
                    st.error(f"Error: {results['error']}")
                else:
                    st.header("Analysis Results:")
                    st.dataframe(results['comparison_df'])

                    st.subheader("Presence Analysis")
                    st.write(f"Number of 'Present in PBI' entries: {results['presence_analysis']['pbi_present_count']}")
                    st.write(f"Number of 'Present in Cognos' entries: {results['presence_analysis']['cognos_present_count']}")

                    if results['presence_analysis']['pbi_present_ids']:
                        st.write("Unique IDs 'Present in PBI':")
                        st.write(results['presence_analysis']['pbi_present_ids'])
                    else:
                        st.write("No Unique IDs 'Present in PBI'.")

                    if results['presence_analysis']['cognos_present_ids']:
                        st.write("Unique IDs 'Present in Cognos':")
                        st.write(results['presence_analysis']['cognos_present_ids'])
                        st.title("🧩 Unique Key Segment Comparator")

                                # Input fields
                        cognos_key = st.text_input("Enter Cognos Unique Key")
                        pbi_key = st.text_input("Enter PBI Unique Key")

                        def extract_mismatched_segment(cognos_key, pbi_key):
                            cognos_parts = cognos_key.split('-')
                            pbi_parts = pbi_key.split('-')

                            mismatch_info = []

                            min_len = min(len(cognos_parts), len(pbi_parts))

                            for i in range(min_len):
                                if cognos_parts[i] != pbi_parts[i]:
                                    column_name = selected_columns[i] if i < len(selected_columns) else f'Segment {i+1}'
                                    mismatch_info.append({
                                        'segment_position': i + 1,
                                        'cognos_value': cognos_parts[i],
                                        'pbi_value': pbi_parts[i],
                                        'length_mismatch': len(cognos_parts[i]) != len(pbi_parts[i]),
                                        'column_name': column_name  # Get the column name from selected_columns
                                    })
                            
                            # Check if one key has more segments than the other
                            if len(cognos_parts) > len(pbi_parts):
                                for i in range(min_len, len(cognos_parts)):
                                    column_name = selected_columns[i] if i < len(selected_columns) else f'Segment {i+1}'
                                    mismatch_info.append({
                                        'segment_position': i + 1,
                                        'cognos_value': cognos_parts[i],
                                        'pbi_value': '[MISSING]',
                                        'length_mismatch': True,
                                        'column_name': column_name # Get the column name
                                    })
                            elif len(pbi_parts) > len(cognos_parts):
                                for i in range(min_len, len(pbi_parts)):
                                    column_name = selected_columns[i] if i < len(selected_columns) else f'Segment {i+1}'
                                    mismatch_info.append({
                                        'segment_position': i + 1,
                                        'cognos_value': '[MISSING]',
                                        'pbi_value': pbi_parts[i],
                                        'length_mismatch': True,
                                        'column_name': column_name # Get the column name
                                    })
                            
                            return mismatch_info
                        # Run comparison if both inputs are provided
                        if cognos_key and pbi_key:
                            mismatches = extract_mismatched_segment(cognos_key, pbi_key)

                            if mismatches:
                                st.write("### ❗ Mismatched Segments Detected:")
                                for mismatch in mismatches:
                                    st.markdown(f"""
                                    🔸 **Segment #{mismatch['segment_position']}**  
                                    - Column Orignial Name #{mismatch['column_name']}
                                    - Cognos: `{mismatch['cognos_value']}`  
                                    - PBI: `{mismatch['pbi_value']}`  
                                    - {'❗ Length mismatch' if mismatch['length_mismatch'] else '⚠️ Value mismatch'}
                                    """)
                                st.warning("These differences might be causing validation issues.")
                            else:
                                st.success("✅ All segments match. No mismatches detected.")
    elif option == "Excel Combiner":
        st.title("🔗Excel Workbook Combiner✨")
        st.write("This app allows you to upload multiple Excel workbooks, rename the sheets, and combine them into one.")

        # --- Helper Function for Sanitizing Sheet Names ---
        def sanitize_sheet_name(name):
            """Removes invalid characters and truncates sheet names for Excel."""
            # Remove invalid characters: \ / * ? : [ ]
            name = re.sub(r'[\\/*?:\[\]]', '_', name)
            # Truncate to Excel's 31-character limit
            return name[:31]

        # --- Helper Function to Copy Sheet with Images ---
        def copy_sheet_with_images(source_sheet, target_workbook, new_sheet_name):
            """Copy a worksheet including images, charts and all cell styles."""
            # Create a new sheet in the target workbook
            target_sheet = target_workbook.create_sheet(title=new_sheet_name)
            
            # Copy cell values, styles, dimensions
            for row in source_sheet.rows:
                for cell in row:
                    new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
            
            # Copy column dimensions
            for key, dimension in source_sheet.column_dimensions.items():
                target_sheet.column_dimensions[key].width = dimension.width
                target_sheet.column_dimensions[key].hidden = dimension.hidden
            
            # Copy row dimensions
            for key, dimension in source_sheet.row_dimensions.items():
                target_sheet.row_dimensions[key].height = dimension.height
                target_sheet.row_dimensions[key].hidden = dimension.hidden
            
            # Copy merged cells
            for merged_cell_range in source_sheet.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_cell_range))
            
            # Copy images and charts
            if source_sheet._images:
                for image in source_sheet._images:
                    target_sheet.add_image(copy(image))
            
            # Handle other drawing objects
            if source_sheet._charts:
                for chart in source_sheet._charts:
                    target_sheet.add_chart(copy(chart))
            
            return target_sheet

        # --- Streamlit UI ---
        # Add file naming inputs in a horizontal layout
        col1, col2 = st.columns(2)
            
        with col1:
            model_name = st.text_input("Model Name", "")
            
        with col2:
            report_name = st.text_input("Report Name", "")

        num_pages = st.number_input("How many Excel workbooks do you want to process?", min_value=1, value=2, step=1)

        uploaded_files = []
        page_names = []

        cols = st.columns(num_pages)

        for i in range(num_pages):
            with cols[i]:
                st.subheader(f"Workbook {i+1}")
                # Use a more descriptive default page name if desired
                default_page_name = f"Source_{i+1}"
                page_name = st.text_input(f"Suffix for sheets from Workbook {i+1}", value=default_page_name, key=f"page_name_{i}")
                uploaded_file = st.file_uploader(f"Upload Excel workbook {i+1}", type=["xlsx", "xls"], key=f"file_{i}")

                uploaded_files.append(uploaded_file)
                page_names.append(page_name)

        # --- Processing Logic ---
        if st.button("Process and Combine Workbooks"):
            # Check if all files are uploaded
            if None in uploaded_files:
                st.error("Please upload all required Excel workbooks.")
            else:
                progress_bar = st.progress(0)
                status_text = st.empty()
                processed_sheets_count = 0
                total_sheets_estimate = 0 # Estimate total sheets for progress bar

                try:
                    # --- Step 1: Estimate total sheets for progress bar ---
                    status_text.text("Analyzing input files...")
                    temp_total_sheets = 0
                    for i, uploaded_file in enumerate(uploaded_files):
                        # Need to reset file pointer after reading names
                        uploaded_file.seek(0)
                        try:
                            xls_temp = pd.ExcelFile(uploaded_file)
                            temp_total_sheets += len(xls_temp.sheet_names)
                        except Exception as e:
                            st.warning(f"Could not read sheet names from Workbook {i+1}. Skipping estimation for this file. Error: {e}")
                        finally:
                            uploaded_file.seek(0) # IMPORTANT: Reset pointer again for actual processing
                    total_sheets_estimate = temp_total_sheets if temp_total_sheets > 0 else 1 # Avoid division by zero

                    # --- Step 2: Create the combined workbook ---
                    status_text.text("Creating combined workbook structure...")
                    combined_wb = openpyxl.Workbook()
                    # Remove the default sheet created by openpyxl
                    if "Sheet" in combined_wb.sheetnames:
                        default_sheet = combined_wb["Sheet"]
                        combined_wb.remove(default_sheet)

                    # Keep track of sheet names used in the *new* workbook to avoid duplicates
                    final_sheet_names_in_workbook = set()
                    all_sheet_names_added = [] # List to show the user which sheets were added

                    # --- Step 3: Process each uploaded workbook ---
                    for i, (uploaded_file, page_name) in enumerate(zip(uploaded_files, page_names)):
                        status_text.text(f"Processing Workbook {i+1} ('{page_name}')...")
                        uploaded_file.seek(0) # Ensure file pointer is at the beginning

                        try:
                            # Save the uploaded file to a temporary location and use openpyxl to open it
                            temp_data = uploaded_file.read()
                            with io.BytesIO(temp_data) as temp_file:
                                # Load workbook with openpyxl
                                source_wb = openpyxl.load_workbook(temp_file, data_only=False)
                                sheet_names = source_wb.sheetnames
                                
                                # Process each sheet in the current workbook
                                for sheet_index, sheet_name in enumerate(sheet_names):
                                    current_sheet_progress = (processed_sheets_count + 1) / total_sheets_estimate
                                    progress_bar.progress(min(current_sheet_progress, 1.0)) # Cap progress at 1.0
                                    status_text.text(f"Processing Workbook {i+1} ('{page_name}') - Sheet: '{sheet_name}'...")
                                    
                                    source_sheet = source_wb[sheet_name]
                                    
                                    # Skip empty sheets
                                    if source_sheet.max_row <= 1 and source_sheet.max_column <= 1:
                                        # Check if only cell A1 exists and is empty
                                        if source_sheet.max_row == 1 and source_sheet.max_column == 1:
                                            if source_sheet.cell(row=1, column=1).value is None and not source_sheet._images:
                                                st.info(f"Skipping empty sheet: '{sheet_name}' from Workbook {i+1}")
                                                processed_sheets_count += 1
                                                continue
                                        else:
                                            st.info(f"Skipping empty sheet: '{sheet_name}' from Workbook {i+1}")
                                            processed_sheets_count += 1
                                            continue

                                    # --- Create and sanitize the new sheet name ---
                                    base_new_sheet_name = f"{sheet_name}_{page_name}"
                                    sanitized_base_name = sanitize_sheet_name(base_new_sheet_name)

                                    # Ensure uniqueness *after* sanitization/truncation
                                    final_sheet_name = sanitized_base_name
                                    counter = 1
                                    while final_sheet_name in final_sheet_names_in_workbook:
                                        suffix = f"_{counter}"
                                        # Ensure the base name + suffix doesn't exceed 31 chars
                                        truncate_at = 31 - len(suffix)
                                        if truncate_at <= 0:
                                            # Handle edge case where suffix itself makes it too long (should be rare)
                                            final_sheet_name = f"Sheet_{processed_sheets_count+1}"[:31] # Fallback name
                                        else:
                                            final_sheet_name = sanitized_base_name[:truncate_at] + suffix
                                        counter += 1
                                        if counter > 100: # Safety break to prevent infinite loops
                                            st.warning(f"Could not generate unique name for sheet derived from '{sheet_name}'/'{page_name}'. Using fallback.")
                                            final_sheet_name = f"Sheet_{processed_sheets_count+1}"[:31]
                                            while final_sheet_name in final_sheet_names_in_workbook:
                                                processed_sheets_count +=1 # Just ensure uniqueness
                                                final_sheet_name = f"Sheet_{processed_sheets_count+1}"[:31]
                                            break # Exit inner while loop

                                    # --- Copy sheet to combined workbook with images ---
                                    copy_sheet_with_images(source_sheet, combined_wb, final_sheet_name)
                                    final_sheet_names_in_workbook.add(final_sheet_name)
                                    all_sheet_names_added.append(final_sheet_name)
                                    processed_sheets_count += 1

                        except Exception as sheet_error:
                            st.warning(f"Could not process Workbook {i+1} ('{page_name}'). Error: {sheet_error}. Skipping this workbook.")
                            st.code(traceback.format_exc())
                            continue # Move to the next workbook

                    progress_bar.progress(1.0) # Ensure progress bar reaches 100%
                    status_text.text("Combining complete. Preparing download...")

                    # --- Step 4: Save and provide download ---
                    if not all_sheet_names_added:
                        st.warning("No data sheets were found or processed in the uploaded workbooks.")
                        status_text.text("") # Clear status
                        progress_bar.empty() # Remove progress bar

                    else:
                        with io.BytesIO() as output:
                            combined_wb.save(output)
                            output.seek(0)
                            data = output.getvalue()

                        st.success("All processable workbooks and sheets combined successfully!")
                        st.write(f"Combined workbook contains the following sheets: {', '.join(sorted(all_sheet_names_added))}")

                        # Generate dynamic filename with current date
                        current_date = datetime.now().strftime("%d-%m-%Y")
                        
                        # Clean model and report names (remove spaces or special characters)
                        clean_model_name = re.sub(r'[\\/*?:"<>|\s]', ' ', model_name) if model_name else "Model"
                        clean_report_name = re.sub(r'[\\/*?:"<>|\s]', ' ', report_name) if report_name else "Report"
                        
                        dynamic_filename = f"{clean_model_name}-{clean_report_name}-{current_date}.xlsx"

                        st.download_button(
                            label="Download Combined Workbook",
                            data=data,
                            file_name=dynamic_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        status_text.text("") # Clear status text on success
                        progress_bar.empty() # Remove progress bar

                except Exception as e:
                    st.error(f"An unexpected error occurred during processing: {e}")
                    st.error("Please check your input files. If the problem persists, check the logs or report the error.")
                    st.code(traceback.format_exc()) # Show detailed error traceback for debugging
                    status_text.text("Processing failed.")
                    progress_bar.empty() # Remove progress bar
                        

if __name__ == "__main__":
    main()
